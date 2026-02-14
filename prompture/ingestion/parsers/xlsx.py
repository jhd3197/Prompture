"""XLSX parser wrapping Tukuy XLSX transformers."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from ..document import DocumentContent
from .base import BaseParser

logger = logging.getLogger("prompture.ingestion.parsers.xlsx")


class XlsxParser(BaseParser):
    """Parse XLSX files using Tukuy's ``XlsxToJsonTransformer``.

    Requires ``openpyxl``.
    Install with ``pip install prompture[ingest]``.
    """

    extensions = [".xlsx", ".xls"]
    mime_types = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]

    def parse(
        self,
        source: str | Path,
        *,
        pages: list[int] | None = None,
        max_file_size: int = 50 * 1024 * 1024,
        sheet: str | None = None,
        header_row: int = 1,
        **kwargs: Any,
    ) -> DocumentContent:
        p = self._check_file(source, max_file_size)
        path_str = str(p)

        # --- Try Tukuy XlsxToJsonTransformer ---
        try:
            from tukuy.plugins.xlsx import XlsxToJsonTransformer

            transformer = XlsxToJsonTransformer("_ingest_xlsx", sheet=sheet, header_row=header_row)
            text = transformer._transform(path_str)

            # Check for error response
            if isinstance(text, str) and text.startswith('{"error"'):
                raise ImportError(text)
        except ImportError:
            # Fallback: openpyxl directly
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise ImportError(
                    "openpyxl is required for XLSX ingestion. Install with: pip install prompture[ingest]"
                ) from None

            text = self._parse_with_openpyxl(path_str, sheet=sheet, header_row=header_row)

        # Parse the JSON text to extract metadata
        metadata: dict[str, Any] = {}
        try:
            data = json.loads(text)
            if isinstance(data, list):
                metadata["row_count"] = len(data)
                if data:
                    metadata["headers"] = list(data[0].keys())
                    metadata["column_count"] = len(data[0])
        except (json.JSONDecodeError, TypeError):
            pass

        # Add sheet info
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path_str, read_only=True)
            metadata["sheets"] = wb.sheetnames
            metadata["active_sheet"] = wb.active.title if wb.active else None
            wb.close()
        except Exception:
            pass

        return DocumentContent(
            text=text,
            file_type="xlsx",
            source_path=path_str,
            metadata=metadata,
            page_count=None,
            page_texts=None,
            char_count=len(text),
        )

    # ------------------------------------------------------------------
    # Fallback
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_with_openpyxl(
        path: str,
        *,
        sheet: str | None = None,
        header_row: int = 1,
    ) -> str:
        """Direct openpyxl parsing without Tukuy."""
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return json.dumps([])

        header_idx = header_row - 1
        if header_idx >= len(rows):
            return json.dumps([])

        headers = [str(h or f"col_{i}") for i, h in enumerate(rows[header_idx])]
        data = []
        for row in rows[header_idx + 1 :]:
            row_dict: dict[str, Any] = {}
            for i, header in enumerate(headers):
                cell = row[i] if i < len(row) else None
                if cell is None:
                    row_dict[header] = None
                elif isinstance(cell, (int, float, bool)):
                    row_dict[header] = cell
                else:
                    row_dict[header] = str(cell)
            data.append(row_dict)

        return json.dumps(data, ensure_ascii=False, default=str)
