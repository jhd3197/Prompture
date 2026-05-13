"""Excel (.xlsx) document loader (openpyxl)."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from ..documents import Document, DocumentLoader, LoaderSource

_XLSX_HINT = "XLSXLoader requires 'openpyxl'. Install with: pip install 'prompture[rag-xlsx]'"


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore
    except ImportError as e:  # pragma: no cover
        raise ImportError(_XLSX_HINT) from e
    return openpyxl


def _row_to_text(row: tuple) -> str:
    return "\t".join("" if cell is None else str(cell) for cell in row)


class XLSXLoader(DocumentLoader):
    """Load an Excel workbook by sheet or by row."""

    supported_extensions = (".xlsx",)

    def load(self, source: LoaderSource, **options: Any) -> list[Document]:
        openpyxl = _require_openpyxl()
        mode = options.get("mode", "sheets")  # "sheets" | "rows"

        if isinstance(source, bytes):
            wb = openpyxl.load_workbook(BytesIO(source), data_only=True)
            source_label = "<bytes>"
        else:
            path = Path(source)
            wb = openpyxl.load_workbook(str(path), data_only=True)
            source_label = str(path)

        docs: list[Document] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if mode == "rows":
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    content = _row_to_text(row)
                    if not content.strip():
                        continue
                    docs.append(
                        Document(
                            content=content,
                            metadata={
                                "source": source_label,
                                "sheet_name": sheet_name,
                                "row_number": i,
                                "type": "xlsx",
                            },
                        )
                    )
            else:
                lines = [_row_to_text(row) for row in ws.iter_rows(values_only=True)]
                content = "\n".join(lines)
                docs.append(
                    Document(
                        content=content,
                        metadata={
                            "source": source_label,
                            "sheet_name": sheet_name,
                            "type": "xlsx",
                        },
                    )
                )
        return docs
